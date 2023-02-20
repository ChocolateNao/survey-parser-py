import openpyxl
# import dotenv
import config


# DO NOT TOUCH
def load_init_sheet_by_id(sheet_index: int):
    if sheet_index < 0:
        raise ValueError("Sheet ID can not be negative")
    init_wb_obj = openpyxl.load_workbook(config.WORKBOOK_DIR + config.WORKBOOK_NAME_INIT)
    init_wb_obj.active = sheet_index
    return init_wb_obj.active


def load_target_sheet_by_id(sheet_index: int):
    if sheet_index < 0:
        raise ValueError("Sheet ID can not be negative")
    target_wb_obj = openpyxl.load_workbook(config.WORKBOOK_DIR + config.WORKBOOK_NAME_TARGET)
    target_wb_obj.active = sheet_index
    return target_wb_obj.active


def init_sheet_max_row(sheet_index: int):
    if sheet_index < 0:
        raise ValueError("Sheet ID can not be negative")
    sheet = load_init_sheet_by_id(sheet_index)
    return sheet.max_row


def init_sheet_max_column(sheet_index: int):
    if sheet_index < 0:
        raise ValueError("Sheet ID can not be negative")
    sheet = load_init_sheet_by_id(sheet_index)
    return sheet.max_column


def target_sheet_max_row(sheet_index: int):
    if sheet_index < 0:
        raise ValueError("Sheet ID can not be negative")
    sheet = load_target_sheet_by_id(sheet_index)
    return sheet.max_row


def target_sheet_max_column(sheet_index: int):
    if sheet_index < 0:
        raise ValueError("Sheet ID can not be negative")
    sheet = load_target_sheet_by_id(sheet_index)
    return sheet.max_column


SHEET_TEACHERS_HEADER = ['Факультет', 'Курс', 'Группа', 'ФИО преподавателя', 'Дисциплина', 'Тип', 'Количество ответов']
SHEET_SUBJECTS_HEADER = ['Факультет', 'Курс', 'Группа', 'ФИО преподавателя', 'Дисциплина', 'Количество ответов']

OPTIONAL_RESPONSE_INFO = {
    'да, в начале семестра': 0,
    'да, перед промежуточной аттестацией': 0,
    'нет': 0
}

OPTIONAL_COURSE_TIME_SPENT = {
    '0-2': 0,
    '2-4': 0,
    '4-6': 0,
    '6-8': 0,
    '>10': 0
}

TEACHER_EVALUATION_CRITERIA = [
    'Общая удовлетворенность преподаванием',
    'Уровень владения преподаваемым материалом',
    'Планирование и организация дисциплины',
    'Коммуникация (доступность преподавателя лично и по почте)',
    'Четкость и структурированность изложения материала',
    'Вовлечение студентов в изучение дисциплины',
    'Комфортная атмосфера на занятиях',
    'Использование цифровых инструментов преподавателем'
]

SUBJECT_EVALUATION_CRITERIA = [

]

SPECIAL_QUESTIONS = {
    'specify_teacher': 'Укажите, пожалуйста, фамилию/и преподавателя/ей',
    'additional_teacher': 'Работал ли с вашей группой по данной дисциплине еще один преподаватель/ментор?'
}
