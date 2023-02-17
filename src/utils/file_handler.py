import os.path
import openpyxl

import config


def ensure_file_workbook(directory: str, filename: str):
    """
    :param directory: Path to the .xlsx file
    :param filename: Name of the .xlsx file
    :return:
    """
    try:
        if os.path.exists(directory):
            if os.path.isfile(os.path.join(directory, filename + '' if filename.endswith('.xlsx') else filename + '.xlsx')):
                print(f"File {config.WORKBOOK_NAME_TARGET} exists! Let's roll!")
            else:
                create_target_workbook(directory, filename)
                print("The specified file cannot be found so it was created.")
        else:
            raise Exception("There was a problem related to the file path.")
    except Exception as e:
        print("An error occurred:", e)


def create_target_workbook(directory: str, filename: str):
    """
    :param directory: Path to the .xlsx file
    :param filename: Name of the .xlsx file
    :return:
    """
    try:
        if os.path.exists(directory):
            target_workbook = openpyxl.Workbook()

            target_workbook.create_sheet(config.TARGET_SHEET_SUBJECTS_NAME)
            target_workbook.create_sheet(config.TARGET_SHEET_TEACHERS_NAME)

            # Deleting the automatically created sheet
            sheet_to_del = target_workbook["Sheet"]
            target_workbook.remove(sheet_to_del)

            target_workbook.save(os.path.join(directory, filename + '' if filename.endswith('.xlsx') else filename + '.xlsx'))
    except Exception as e:
        print("An error occurred:", e)
