import os
from typing import List, Any, Tuple

import openpyxl
from openpyxl.utils import get_column_letter

import config
from sheet_processing import get_teachers, get_header_data, \
    get_dictionary_by_teacher, get_subjects_dict, get_subjects
from utils.constants import init_sheet_max_row, load_init_sheet_by_id, optional_course_time_spent, \
    optional_response_info, load_target_sheet_by_id, target_sheet_max_row
from utils.regex import *
from utils.utils import to_fixed, replace_strings_with_zero


def load_teacher_feedback(teacher_dict: dict, teacher_name: str) -> tuple[
                                                                        list[int | str | Any], list[int | str | Any]] | \
                                                                        list[int | str | Any] | None:
    """
    :param teacher_name: A SINGLE key = teacher name
    :param teacher_dict: get_dictionary_by_teacher(get_header_data(), get_teachers())
    :return: A list of format [Faculty, Year, Group, Initials, Subject, Type, Recipients count, Grades..., Comments]
    """
    # INIT
    sheet_init = load_init_sheet_by_id(0)
    max_row_init = init_sheet_max_row(0)

    base_len = 10
    data_row = []
    teacher_list = teacher_dict[teacher_name]

    for feedback in teacher_list:
        grades_header = []

        for row in range(max_row_init + 1):
            if row >= 1:
                cell_init = sheet_init.cell(row=row, column=feedback[1])
                if cell_init.value is not None:
                    grades_header.append(cell_init.value)

        grades_data = grades_header[1:]

        participants_count = len(grades_data)
        if participants_count not in data_row:
            data_row.insert(0, participants_count)
        if len(data_row) > base_len and participants_count not in data_row[10:]:
            data_row.insert(10, participants_count)

        if len(grades_data) >= 1 and grades_data[0] is not None:
            if type(grades_data[0]) != str:
                grade_mean = sum(grades_data) / len(grades_data)
                if grade_mean >= 0:
                    avg_grade = to_fixed(grade_mean, 2)
                    data_row.append(avg_grade)
                else:
                    data_row.append(0)

            if type(grades_data[0]) == str:
                comments = '; '
                comments = comments.join(grades_data)
                data_row.append(comments.strip())

    if len(data_row) > base_len:
        if len(data_row) == base_len * 2:
            mid = len(data_row) // 2
            data_tuple = data_row[:mid], data_row[mid:]

            for data_rrow in data_tuple:
                config_data_row = [config.TARGET_CELL_GROUP, config.TARGET_CELL_GRADE, config.TARGET_CELL_FACULTY]

                for config_row in config_data_row:
                    data_rrow.insert(0, config_row)
                find_name = regex_name_find(grades_header[0])
                subjects = []
            for item in teacher_list:
                find_subject = regex_parenthesis_find(item[0])
                if find_subject[0] not in subjects:
                    subjects.append(find_subject[0])
            subjects.pop(1)
            subjects.pop(2)

            data_tuple[0].insert(3, find_name[0])
            data_tuple[1].insert(3, find_name[0])
            data_tuple[0].insert(4, subjects[0][1:][:-1])
            data_tuple[1].insert(4, subjects[1][1:][:-1])

            return data_tuple

    # Always present
    config_data_row = [config.TARGET_CELL_GROUP, config.TARGET_CELL_GRADE, config.TARGET_CELL_FACULTY]
    for config_row in config_data_row:
        data_row.insert(0, config_row)

    find_subject = regex_parenthesis_find(teacher_list[0][0])
    find_name = regex_name_find(teacher_list[0][0])
    data_row.insert(3, find_name[0])
    data_row.insert(4, find_subject[0][1:][:-1])

    # Optional
    find_type = regex_class_type_find(teacher_list[0][0])
    if find_type is not None:
        data_row.insert(5, find_type.strip())
    else:
        data_row.insert(5, '')

    if data_row[-1] != 0:
        return data_row  # [Faculty, Year, Group, Initials, Subject, Type, Recipients count, Grades..., Comments]
    else:
        return  # Не бейте за эту функцию, дело было в 4 часа утра, когда-нибудь перепишу :(


def write_teacher_data():
    # TARGET
    workbook_path_init = os.path.join(config.WORKBOOK_DIR, config.WORKBOOK_NAME_TARGET)
    workbook_init = openpyxl.load_workbook(workbook_path_init)
    sheet_init = workbook_init[config.TARGET_SHEET_TEACHERS_NAME]

    teachers = get_teachers(counter=True)
    row = sheet_init.max_row + 1
    header_data = get_header_data()
    teachers_data = get_teachers()
    teacher_dict = get_dictionary_by_teacher(header_data, teachers_data)
    for teacher in teachers_data:
        func = load_teacher_feedback(teacher_dict, teacher)
        if func is not None and type(func) == list:

            for col, col_name in enumerate(func, start=1):
                cell = sheet_init.cell(row=row, column=col)
                cell.value = col_name

                sheet_init.column_dimensions[get_column_letter(col)].width = 20
                if col == len(func):
                    row += 1

        if func is not None and type(func) == tuple:

            for tuple_data in func:
                tuple_data.insert(5, '')
                for col, col_name in enumerate(tuple_data, start=1):
                    cell = sheet_init.cell(row=row, column=col)
                    cell.value = col_name

                    sheet_init.column_dimensions[get_column_letter(col)].width = 20
                    if col == len(tuple_data):
                        row += 1

    workbook_init.save(workbook_path_init)


def get_subject_feedback(subject_dict: dict, subject_name: str):
    # INIT
    sheet_init = load_init_sheet_by_id(0)
    max_row_init = init_sheet_max_row(0)

    data_row = []

    subject_dict = subject_dict[subject_name]

    for feedback in subject_dict:
        grades_header = []

        for row in range(max_row_init + 1):
            if row >= 1:
                cell_init = sheet_init.cell(row=row, column=feedback[1])
                if cell_init.value is not None:
                    grades_header.append(cell_init.value)

        grades_data = grades_header[1:]
        for grade, i in enumerate(grades_data):
            if i == 'нет домашних заданий' or grade == 'нет домашних заданий':
                grades_data = replace_strings_with_zero(grades_data)

        participants_count = len(grades_data)
        if participants_count not in data_row:
            data_row.insert(0, participants_count)

        if len(grades_data) >= 1 and grades_data[0] is not None:
            if type(grades_data[0]) != str:
                grade_mean = sum(grades_data) / len(grades_data)
                if grade_mean >= 0:
                    avg_grade = to_fixed(grade_mean, 2)
                    data_row.append(avg_grade)
                else:
                    data_row.append(0)
            if grades_data[0] == 'да, в начале семестра' or grades_data[0] == 'да, перед промежуточной аттестацией' or \
                    grades_data[0] == 'нет':
                response_tuple_1 = optional_response_info

                for grade in grades_data:
                    for response in response_tuple_1:
                        if grade == response:
                            response_tuple_1[response] += 1
                data_row.append(str(response_tuple_1).strip('{}'))

            if grades_data[0] == '0-2' or grades_data[0] == '2-4' or grades_data[0] == '4-6' or grades_data[
                0] == '6-8' or grades_data[0] == '>10':
                response_tuple_2 = optional_course_time_spent

                for grade in grades_data:
                    for response in response_tuple_2:
                        if grade == response:
                            response_tuple_2[response] += 1
                data_row.append(str(response_tuple_2).strip('{}'))

            if type(grades_data[0]) == str and grades_data[0] != 'да, в начале семестра' and grades_data[
                0] != 'да, перед промежуточной аттестацией' and grades_data[0] != 'нет' and grades_data[0] != '0-2' and \
                    grades_data[0] != '2-4' and grades_data[0] != '4-6' and grades_data[0] != '6-8' and grades_data[
                0] != '>10' and grades_data[0] != 'нет домашних заданий':
                comments = '; '
                comments = comments.join(grades_data)
                data_row.append(comments.strip())

    # Always present
    config_data_row = [config.TARGET_CELL_GROUP, config.TARGET_CELL_GRADE, config.TARGET_CELL_FACULTY]
    for config_row in config_data_row:
        data_row.insert(0, config_row)

    find_subject = regex_braces_find(subject_dict[0][0])
    data_row.insert(3, find_subject[0][1:][:-1])

    data_row.insert(2, '')

    # Optional

    if data_row[-1] != 0:
        return data_row  # [Faculty, Year, Group, Initials, Subject, Recipients count, Grades..., Comments]
    else:
        return  # Не бейте за эту функцию, дело было в 4 часа утра, когда-нибудь перепишу :(


def write_subject_data():
    # TARGET
    workbook_path_init = os.path.join(config.WORKBOOK_DIR, config.WORKBOOK_NAME_TARGET)
    workbook_init = openpyxl.load_workbook(workbook_path_init)
    sheet_init = workbook_init[config.TARGET_SHEET_SUBJECTS_NAME]

    row = sheet_init.max_row + 1
    subjects_data = get_subjects()
    subjects_dict = get_subjects_dict()
    for subject in subjects_data:
        func = get_subject_feedback(subjects_dict, subject)
        if func is not None and type(func) == list:

            for col, col_name in enumerate(func, start=1):
                cell = sheet_init.cell(row=row, column=col)
                cell.value = col_name

                sheet_init.column_dimensions[get_column_letter(col)].width = 20
                if col == len(func):
                    row += 1

    workbook_init.save(workbook_path_init)


# def fetch_teachers():
#     sheet_target = load_target_sheet_by_id(1)  # Teachers - id 1, Subjects - id 0
#     teachers_column = 3
#     subjects_column = 4
#     max_row = target_sheet_max_row(1)
#     subjects_key = []
#     dictionary =
#
#     for row in range(max_row):
#         cell = sheet_target.cell(row=row, column=subjects_column)
#         cell_value = cell.value
#         if cell_value not in subjects_key:  # Subject - key
#             subjects_key.append(cell_value)
#
#     for row in range(max_row):
#         cell = sheet_target.cell(row=row, column=subjects_column)
#         cell_value = cell.value




