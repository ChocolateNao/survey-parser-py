import os

import openpyxl

import config
from sheet_processing import get_dictionary_by_subject, get_teachers, get_subjects, get_header_data, get_subjects_dict
from utils.constants import const_sheet_teachers_header, const_sheet_subjects_header
from utils.regex import regex_braces_find, regex_braces_remove, regex_remove_text_in_braces


def get_subject_sheet_header() -> list:
    questions_braces = []

    subjects_data = get_subjects()

    dictionary = get_subjects_dict()
    dictionary = dictionary[subjects_data[2]]

    for item in dictionary:
        regex = regex_remove_text_in_braces(item[0])
        if subjects_data[2] in item[0] and regex not in questions_braces:
            questions_braces.append(regex.strip())

    # questions = []
    # for question in questions_braces:
    #     question = regex_braces_remove(question[0])
    #     questions.append(question)
    # print(questions)
    return questions_braces


def insert_target_subjects_header():
    try:
        workbook_path = os.path.join(config.WORKBOOK_DIR, config.WORKBOOK_NAME_TARGET)
        workbook = openpyxl.load_workbook(workbook_path)
        worksheet_subjects = workbook[config.TARGET_SHEET_SUBJECTS_NAME]

        header = get_subject_sheet_header()
        header = const_sheet_subjects_header + header[:-1] + ['Комментарии']

        for i, col_name in enumerate(header, start=1):
            cell = worksheet_subjects.cell(row=1, column=i)
            cell.value = col_name
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

            # if i == 7:
            #     column1 = i  # column index for A
            #     column2 = i + 3
            #     cell_range = get_column_letter(column1) + str(1) + ":" + get_column_letter(column2) + str(1)
            #     worksheet_subjects.merge_cells(f'{get_column_letter(i)}:{get_column_letter(i + 3)}')
            #     i = i + 3

        workbook.save(os.path.join(workbook_path))
    except PermissionError as e:
        print(f"Permission error while writing data in '{workbook_path}'! The current workbook may be opened in "
              f"another editor\n{e}")


# Факультет
# Курс
# Группа
# ФИО преподавателя
# Дисциплина
# Количество ответов
# Удовлетворены ли Вы реализацией дисциплины? (0 - полностью НЕ удовлетворен, 1-скорее не удовлетворен, 2-скорее удовлетворен, 3-полностью удовлетворен)
# Были ли  вы проинформированы о целях курса, его  месте в образовательной программе и о критериях оценивания?
#     да, в начале семестра
#     да, перед промежуточной аттестацией
#     нет
# Оцените сложность курса (1 - очень легко, 5 - очень сложно)
# Оцените сложность домашних заданий (1 - очень простые, 5 - очень сложные)
# Сколько часов в неделю (кроме аудиторных) Вы тратите на курс?
#     0 - 2
#     2 - 4
#     4 - 6
#     6 - 8 > 10
# Открытые комментарии

def get_teacher_sheet_header() -> list:
    questions_braces = []

    subjects_data = get_subjects()
    teachers_data = get_teachers()
    header_data = get_header_data()

    dictionary = get_dictionary_by_subject(header_data, subjects_data)
    dictionary = dictionary[subjects_data[2]]

    for item in dictionary:
        for teacher in teachers_data:
            regex = regex_braces_find(item[0])
            if teacher in item[0] and regex not in questions_braces:
                questions_braces.append(regex)

    questions = []
    for question in questions_braces:
        question = regex_braces_remove(question[0])
        questions.append(question)

    return questions


def insert_target_teachers_header():
    try:
        workbook_path = os.path.join(config.WORKBOOK_DIR, config.WORKBOOK_NAME_TARGET)
        workbook = openpyxl.load_workbook(workbook_path)
        worksheet_teachers = workbook[config.TARGET_SHEET_TEACHERS_NAME]

        header = get_teacher_sheet_header()
        header = const_sheet_teachers_header + header + ['Комментарии']

        for i, col_name in enumerate(header, start=1):
            cell = worksheet_teachers.cell(row=1, column=i)
            cell.value = col_name
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')

        workbook.save(os.path.join(workbook_path))
    except PermissionError as e:
        print(f"Permission error while writing data in '{workbook_path}'! The current workbook may be opened in "
              f"another editor\n{e}")

# Факультет
# Курс
# Группа
# ФИО преподавателя
# Дисциплина
# Тип
# Количество ответов
# Общая удовлетворенность преподаванием (0 - полностью НЕ удовлетворен, 1 - скорее не удовлетворен, 2 - скорее удовлетворен, 3 - полностью удовлетворен)
# Уровень владения преподаваемым материалом
# Планирование и организация дисциплины
# Коммуникация (доступность преподавателя лично и по почте)
# Четкость и структурированность изложения материала
# Вовлечение студентов в изучение дисциплины
# Комфортная атмосфера на занятиях
# Использование цифровых инструментов преподавателем
# Открытые комментарии
