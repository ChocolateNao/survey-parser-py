import openpyxl
import config


init_wb_obj = openpyxl.load_workbook(config.sheet_dir + config.sheet_name_init)
init_wb_obj.active = 0
init_sheet_obj = init_wb_obj.active

max_row_init_sheet = init_sheet_obj.max_row
max_column_init_sheet = init_sheet_obj.max_column

